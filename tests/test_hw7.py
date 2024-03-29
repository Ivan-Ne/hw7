from zipfile import ZipFile
from pypdf import PdfReader
from openpyxl import load_workbook
from tests.conftest import zip_dir


def test_pdf():
    with ZipFile(zip_dir, "a") as myzip:
        with myzip.open('book.pdf') as book_pdf:
            reader = PdfReader(book_pdf)
            assert 'Python Testing with pytest' in reader.pages[1].extract_text()
            assert myzip.getinfo('book.pdf').file_size == 3035139


def test_xlsx_content():
    with ZipFile(zip_dir, "a") as myzip:
        with myzip.open('table.xlsx') as table_xlsx:
            workbook = load_workbook(table_xlsx)
            sheet = workbook.active
            assert sheet.cell(row=2, column=3).value == 's'
            assert myzip.getinfo('table.xlsx').file_size == 4988


def test_csv_content():
    with ZipFile(zip_dir, "a") as myzip:
        with myzip.open('table.csv') as table_csv:
            sheet = table_csv.read().decode('utf-8-sig')
            assert 'Apple' in sheet
            assert myzip.getinfo('table.csv').file_size == 127
